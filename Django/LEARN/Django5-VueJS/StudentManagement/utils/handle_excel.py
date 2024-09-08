import openpyxl
import datetime
from io import BytesIO


class ReadExcel():
    def __init__(self, file_path):
        self.workbook = openpyxl.load_workbook(file_path)
        self.worksheet = self.workbook.active
        self.data = []

    def get_data(self):
        for row in self.worksheet.iter_rows():
            row_value = []
            for cell in row:
                row_value.append(cell.value)
            self.data.append(row_value)


class WriteExcel():
    # data 可以是 tuple-list 也可以是 list-list
    def __init__(self, file_path, data):
        self.file_path = file_path
        self.data = data
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active

    def write_to_excel(self):
        for row in self.data:
            self.worksheet.append(row)
        self.workbook.save(self.file_path)


class ExportExcel():
    def __init__(self, data):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.data = data

    def export_to_execel(self):
        first_column = ['班级', '姓名', '学号', '性别', '出生日期', '联系电话', '家庭住址']
        self.ws.append(first_column)
        for row in self.data:
            self.ws.append(row)
        excel_file = BytesIO()
        self.wb.save(excel_file)
        self.wb.close()
        return excel_file

# student_data_path = "/Users/Chaksw/Web-Dev/Django课程资料包/StudentManagement/1年1班学生信息.xlsx"
# wb = openpyxl.load_workbook(student_data_path)
# ws = wb.active

# data = []
# for row in ws.iter_rows():
#     # print(row)
#     row_value = []
#     for cell in row:
#         row_value.append(cell.value)
#     data.append(row_value)
# print(data)


if __name__ == '__main__':
    read_excel_object = ReadExcel("/Users/Chaksw/Web-Dev/Django课程资料包/StudentManagement/1年1班学生信息.xlsx")
    read_excel_object.get_data()
    print(read_excel_object.data)
    data = []

    write_excel_object = WriteExcel('学生信息.xlsx', data)
    write_excel_object.write_to_excel()
